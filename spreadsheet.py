import openpyxl

def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    wb = openpyxl.load.workbook(file)
    sheet = wb[sheetName]
    return(sheet.max_column)

def readData(file, sheetName, rownum, columnno):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.cell(row = rownum, column = columnno).value

def writeData(file, sheetName, rownum, columnno, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    wb.save(file)